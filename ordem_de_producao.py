#https://www.youtube.com/watch?v=bu5wXjz2KvU

import pandas as pd
import numpy as np
import os
import datetime 
import gspread
import streamlit as st
import time
import zipfile

from datetime import datetime
from datetime import timedelta
from pathlib import Path
from openpyxl import Workbook, load_workbook
from PIL import Image

###### CONECTANDO PLANILHAS ##########

st.title('Gerador de Ordem de Produção')

st.write("Planilha usada como base para gerar as ordens de produção")
st.write("https://docs.google.com/spreadsheets/d/18ZXL8n47qSLFLVO5tBj7-ADpqmMyFwCgs4cxxtBB9Xo/edit#gid=0")

name_sheet = 'Bases para sequenciamento'

worksheet1 = 'Base_Carretas'
worksheet2 = 'Carga_Vendas'

worksheet3 = 'Base_Carretas'

filename = "service_account.json"

sa = gspread.service_account(filename)
sh = sa.open(name_sheet)

wks1 = sh.worksheet(worksheet1)
wks2 = sh.worksheet(worksheet2)
wks3 = sh.worksheet(worksheet3)

#obtendo todos os valores da planilha
list1 = wks1.get_all_records()
list2 = wks2.get_all_records()

#transformando em dataframe
base_carretas = pd.DataFrame(list1)
base_carga = pd.DataFrame(list2)

###### TRATANDO DADOS #########

#####Tratando datas######

base_carga = base_carga[['PED_PREVISAOEMISSAODOC', '3o. Agrupamento', 'PED_RECURSO.CODIGO', 'PED_QUANTIDADE']]
base_carga['PED_PREVISAOEMISSAODOC'] = pd.to_datetime(base_carga['PED_PREVISAOEMISSAODOC'], format='%d/%m/%Y', errors='coerce')
base_carga['Ano'] = base_carga['PED_PREVISAOEMISSAODOC'].dt.strftime('%Y')
base_carga['PED_PREVISAOEMISSAODOC'] = base_carga.PED_PREVISAOEMISSAODOC.dt.strftime('%d/%m/%Y')

####renomeando colunas#####

base_carga = base_carga.rename(columns={'PED_PREVISAOEMISSAODOC': 'Datas',
                                        '3o. Agrupamento': 'Carga',
                                        'PED_RECURSO.CODIGO': 'Recurso',
                                        'PED_QUANTIDADE':'Qtde'})

#####Valores nulos######

base_carga.dropna(inplace=True)
base_carga.reset_index(drop=True)

today = datetime.now()
ts = pd.Timestamp(today)
today = today.strftime('%d/%m/%Y')

filenames=[]

def unique(list1):
    x = np.array(list1)
    print(np.unique(x))


with st.sidebar:

    image = Image.open('logo-cemagL.png')
    st.image(image, width=300)

with st.form(key='my_form'):
    
    with st.sidebar:
        
        tipo_filtro = st.date_input('Data: ')
        tipo_filtro = tipo_filtro.strftime("%d/%m/%Y")
        #tipo_filtro = "29/07/2022"
        values = ['Selecione','Pintura','Montagem','Solda']
        setor = st.selectbox('Escolha o setor', values)
        
        att_controle = st.selectbox('Atualizar Controle', ['Selecione','Atualizar', 'Não Atualizar'])
        
        att_apontamento = st.selectbox('Atualizar Apontamento', ['Selecione','Atualizar', 'Não Atualizar'])

        submit_button = st.form_submit_button(label='Gerar')

if submit_button:
        
    if setor == 'Pintura':
            
        base_carretas.drop(['Etapa','Etapa3'], axis=1, inplace=True)
        
        base_carretas.drop(base_carretas[(base_carretas['Etapa2']  == '')].index, inplace=True) # & (base_carretas['Unit_Price'] < 600)].index, inplace=True)
        
        base_carretas = base_carretas.reset_index(drop=True)
        
        base_carretas = base_carretas.astype(str)
        
        for d in range(0,base_carretas.shape[0]):
                
            if len(base_carretas['Código'][d]) == 5:
                base_carretas['Código'][d] = '0' + base_carretas['Código'][d]
        
        #separando string por "-" e adicionando no dataframe antigo
        
        tratando_coluna = base_carga["Recurso"].str.split(" - ", n = 1, expand = True)
        
        base_carga['Recurso'] = tratando_coluna[0]
        
        #tratando cores da string
        
        base_carga['Recurso_cor'] = base_carga['Recurso']
        
        base_carga = base_carga.reset_index(drop=True)
        
        df_cores = pd.DataFrame({'Recurso_cor':['AN','VJ','LC','VM','AV','sem_cor'], 
                                 'cor':['Azul','Verde','Laranja','Vermelho','Amarelo','Laranja']})
        
        cores = ['AM','AN','VJ','LC','VM','AV']
        
        base_carga = base_carga.astype(str)
        
        for r in range(0,base_carga.shape[0]):
            base_carga['Recurso_cor'][r] = base_carga['Recurso_cor'][r][len(base_carga['Recurso_cor'][r])-3:len(base_carga['Recurso_cor'][r])]
            base_carga['Recurso_cor'] = base_carga['Recurso_cor'].str.strip()
            
            if len(base_carga['Recurso_cor'][r]) > 2:
                base_carga['Recurso_cor'][r] = base_carga['Recurso_cor'][r][1:3]
                        
            if base_carga['Recurso_cor'][r] not in cores:
                base_carga['Recurso_cor'][r] = "LC"
                
        base_carga = pd.merge(base_carga, df_cores, on=['Recurso_cor'], how='left')
                
        base_carga['Recurso']=base_carga['Recurso'].str.replace('AN','') # Azul
        base_carga['Recurso']=base_carga['Recurso'].str.replace('VJ','') # Verde
        base_carga['Recurso']=base_carga['Recurso'].str.replace('LC','') # Laranja
        base_carga['Recurso']=base_carga['Recurso'].str.replace('VM','') # Vermelho
        base_carga['Recurso']=base_carga['Recurso'].str.replace('AV','') # Amarelo
        
        base_carga['Recurso'] = base_carga['Recurso'].str.strip()
        
        datas_unique = pd.DataFrame(base_carga['Datas'].unique())
            
        escolha_data = (base_carga['Datas'] == tipo_filtro)
        filtro_data = base_carga.loc[escolha_data]
        filtro_data['Datas'] = pd.to_datetime(filtro_data.Datas)
           
        #procv e trazendo as colunas que quero ver
        
        tab_completa = pd.merge(filtro_data, base_carretas, on=['Recurso'], how='left')
        
        tab_completa['Código'] = tab_completa['Código'].astype(str)
        
        tab_completa = tab_completa.reset_index(drop=True)
        
        celulas_unique = pd.DataFrame(tab_completa['Célula'].unique())
        celulas_unique = celulas_unique.dropna(axis=0)
        celulas_unique.reset_index(drop=True)
        
        recurso_unique =  pd.DataFrame(tab_completa['Recurso'].unique())
        recurso_unique = recurso_unique.dropna(axis=0)
        
        #tratando coluna de código
        
        for t in range(0,tab_completa.shape[0]):
            
            if len(tab_completa['Código'][t]) == 5:
                tab_completa['Código'][t] = '0' + tab_completa['Código'][t][0:5]
                
            if len(tab_completa['Código'][t]) == 8:
                tab_completa['Código'][t] = tab_completa['Código'][t][0:6]
                
        #criando coluna de quantidade total de itens
        
        tab_completa['Qtde_x'] = tab_completa['Qtde_x'].str.replace(',','.')
        
        tab_completa['Qtde_x'] = tab_completa['Qtde_x'].astype(float)
        tab_completa['Qtde_x'] = tab_completa['Qtde_x'].astype(int)
        
        tab_completa = tab_completa.dropna(axis=0)
        
        tab_completa['Qtde_y'] = tab_completa['Qtde_y'].astype(float)
        tab_completa['Qtde_y'] = tab_completa['Qtde_y'].astype(int)
        
        tab_completa['Qtde_total'] = tab_completa['Qtde_x'] * tab_completa['Qtde_y']
        
        tab_completa = tab_completa.drop(columns=['Carga','Recurso','Qtde_x','Qtde_y','LEAD TIME','flag peça','Etapa2'])
        
        tab_completa = tab_completa.groupby(['Código','Peca','Célula','Datas','Recurso_cor','cor']).sum()
        tab_completa.reset_index(inplace=True)
        
        tab_completa.drop(tab_completa.loc[tab_completa['Célula']=='EIXO SIMPLES'].index, inplace=True)
        tab_completa.reset_index(inplace=True, drop=True)
        
        for t in range(0,len(tab_completa)):
            
            if tab_completa['Célula'][t] == 'FUEIRO' or \
            tab_completa['Célula'][t] == 'LATERAL' or \
            tab_completa['Célula'][t] == 'PLAT. TANQUE. CAÇAM.': 
                
                tab_completa['Recurso_cor'][t] = tab_completa['Código'][t] + tab_completa['Recurso_cor'][t]
            
            else:
                
                tab_completa['Recurso_cor'][t] = tab_completa['Código'][t] + 'CO' 
                tab_completa['cor'][t] = 'Cinza'
                
        ###########################################################################################
        
        k = 9
        
        cor_unique = tab_completa['cor'].unique()
        
        st.write("Arquivos para download")
        
        for i in range(0,len(cor_unique)):
            
            wb = Workbook()
            wb = load_workbook('modelo_op_pintura.xlsx')
            ws = wb.active
            
            filtro_excel = (tab_completa['cor'] == cor_unique[i])
            filtrar = tab_completa.loc[filtro_excel]
            filtrar = filtrar.reset_index(drop=True)
            filtrar = filtrar.groupby(['Código','Peca','Célula','Datas','Recurso_cor','cor']).sum().reset_index()
            filtrar.sort_values(by=['Célula'], inplace=True)  
            filtrar = filtrar.reset_index(drop=True)
            
            if len(filtrar) > 21:
            
                for j in range(0,21):
                 
                    ws['F5'] = cor_unique[i] # nome da coluna é '0'
                    ws['AD5'] = datetime.now() #  data de hoje
                    ws['M4']  = tipo_filtro  # data da carga
                    ws['B' + str(k)] = filtrar['Recurso_cor'][j]
                    ws['G' + str(k)] = filtrar['Peca'][j]
                    ws['AD' + str(k)] = filtrar['Qtde_total'][j]
                    k = k + 1
                    
                    wb.template = False
                    wb.save("Pintura " + cor_unique[i] +'1.xlsx')
                    
                my_file = "Pintura " + cor_unique[i] +'1.xlsx'
                filenames.append(my_file)                       
                
                k = 9
                
                wb = Workbook()
                wb = load_workbook('modelo_op_pintura.xlsx')
                ws = wb.active
                
                filtro_excel = (tab_completa['cor'] == cor_unique[i])
                filtrar = tab_completa.loc[filtro_excel]
                filtrar = filtrar.reset_index(drop=True)
                filtrar = filtrar.groupby(['Código','Peca','Célula','Datas','Recurso_cor','cor']).sum().reset_index()
                filtrar.sort_values(by=['Célula'], inplace=True)  
                filtrar = filtrar.reset_index(drop=True)
        
                if len(filtrar) > 21:
                    
                    j = 21
                    
                    for j in range(21,len(filtrar)):
                     
                        ws['F5'] = cor_unique[i] # nome da coluna é '0'
                        ws['AD5'] = datetime.now() #  data de hoje
                        ws['M4']  = tipo_filtro  # data da carga
                        ws['B' + str(k)] = filtrar['Recurso_cor'][j]
                        ws['G' + str(k)] = filtrar['Peca'][j]
                        ws['AD' + str(k)] = filtrar['Qtde_total'][j]
                        k = k + 1
                        
                        wb.save("Pintura " + cor_unique[i] +'.xlsx') 
                        
                    my_file = "Pintura " + cor_unique[i] +'.xlsx'
                    filenames.append(my_file)              
            
            else:
                
                j = 0
                k = 9
                for j in range(0,21-(21-len(filtrar))):
                 
                    ws['F5'] = cor_unique[i] # nome da coluna é '0'
                    ws['AD5'] = datetime.now() #  data de hoje
                    ws['M4']  = tipo_filtro  # data da carga
                    ws['B' + str(k)] = filtrar['Recurso_cor'][j]
                    ws['G' + str(k)] = filtrar['Peca'][j]
                    ws['AD' + str(k)] = filtrar['Qtde_total'][j]
                    k = k + 1
                    
                    wb.template = False
                    wb.save("Pintura " + cor_unique[i] +'.xlsx')
                    
                    my_file = "Pintura " + cor_unique[i] +'.xlsx'
                    filenames.append(my_file)                    
                    
                k = 9
                
                my_file = "Pintura " + cor_unique[i] +'.xlsx'
                filenames.append(my_file)
                        
        with st.sidebar:
            
            if att_apontamento == "Atualizar":
                
                tab_completa['Carimbo'] = tipo_filtro + 'Pintura'
                tab_completa['Tinta'] = ''
                tab_completa['Data_carga'] = tipo_filtro
                
                tab_completa1 = tab_completa[['Carimbo','Célula','Recurso_cor','Peca','cor','Tinta','Qtde_total']]
                
                tab_completa1['Data_carga'] = tipo_filtro
                
                tab_completa1 = tab_completa1.astype(str)
                
                name_sheet1 = 'RQ PC-005-002 APONTAMENTO PINTURA M22'
                worksheet3 = 'Sequenciamento automatico -> L'
                
                sa = gspread.service_account(filename)
                sh = sa.open(name_sheet1)
                
                tab_completa1 = tab_completa1.values.tolist()
                
                sh.values_append('Sequenciamento automatico -> L', {'valueInputOption': 'RAW'}, {'values': tab_completa1})
    
    if setor == 'Montagem':
            
        base_carretas['Código'] = base_carretas['Código'].astype(str) 
        base_carretas['Recurso'] = base_carretas['Recurso'].astype(str)
    
        #######retirando cores dos códigos######
    
        base_carga['Recurso']=base_carga['Recurso'].str.replace('AM','')
        base_carga['Recurso']=base_carga['Recurso'].str.replace('AN','')
        base_carga['Recurso']=base_carga['Recurso'].str.replace('VJ','')
        base_carga['Recurso']=base_carga['Recurso'].str.replace('LC','')
        base_carga['Recurso']=base_carga['Recurso'].str.replace('VM','')
        base_carga['Recurso']=base_carga['Recurso'].str.replace('AV','')
    
        ######retirando espaco em branco####
    
        base_carga['Recurso'] = base_carga['Recurso'].str.strip()
    
        #####excluindo colunas e linhas#####
    
        base_carretas.drop(['Etapa2','Etapa3'], axis=1, inplace=True)
    
        base_carretas.drop(base_carretas[(base_carretas['Etapa']  == '')].index, inplace=True) # & (base_carretas['Unit_Price'] < 600)].index, inplace=True)
    
        ####criando código único#####
    
        codigo_unico = tipo_filtro[:2] + tipo_filtro[3:5] + tipo_filtro[6:10]
    
        ####filtrando data da carga#####
    
        datas_unique = pd.DataFrame(base_carga['Datas'].unique())
    
        escolha_data = (base_carga['Datas'] == tipo_filtro)
        filtro_data = base_carga.loc[escolha_data]
        filtro_data['Datas'] = pd.to_datetime(filtro_data.Datas)
    
        #####juntando planilhas de acordo com o recurso#######
    
        tab_completa = pd.merge(filtro_data, base_carretas[['Recurso','Código','Peca','Qtde','Célula']], on=['Recurso'], how='left')
        tab_completa = tab_completa.dropna(axis=0)
    
        tab_completa['Código'] = tab_completa['Código'].astype(str)
    
        tab_completa.reset_index(inplace=True, drop=True)
    
        celulas_unique = pd.DataFrame(tab_completa['Célula'].unique())
        celulas_unique = celulas_unique.dropna(axis=0)
        celulas_unique.reset_index(inplace=True)
    
        recurso_unique =  pd.DataFrame(tab_completa['Recurso'].unique())
        recurso_unique = recurso_unique.dropna(axis=0)
                
        #criando coluna de quantidade total de itens
    
        try:
            tab_completa['Qtde_x'] = tab_completa['Qtde_x'].str.replace(',','.')
        except:
            pass
    
        tab_completa['Qtde_x'] = tab_completa['Qtde_x'].astype(float)
        tab_completa['Qtde_x'] = tab_completa['Qtde_x'].astype(int)
    
        tab_completa['Qtde_y'] = tab_completa['Qtde_y'].astype(float)
        tab_completa['Qtde_y'] = tab_completa['Qtde_y'].astype(int)
    
        tab_completa['Qtde_total'] = tab_completa['Qtde_x'] * tab_completa['Qtde_y']
    
        tab_completa = tab_completa.drop(columns=['Carga','Recurso','Qtde_x','Qtde_y'])
    
        tab_completa = tab_completa.groupby(['Código','Peca','Célula','Datas']).sum()
    
        #tab_completa = tab_completa.drop_duplicates()
    
        tab_completa.reset_index(inplace=True)
    
        ######tratando coluna de código e recurso
    
        for d in range(0,tab_completa.shape[0]):
    
            if len(tab_completa['Código'][d]) == 5:
                tab_completa['Código'][d] = '0' + tab_completa['Código'][d]
                
        #criando coluna de código para arquivar
        
        hoje = datetime.now()  
        
        ts = pd.Timestamp(hoje)
        
        hoje1 = hoje.strftime('%d%m%Y')
        
        controle_seq = tab_completa
        controle_seq["codigo"] = hoje1 + tipo_filtro    
    
        st.write("Arquivos para download")
        
        k = 9

        for i in range(0,len(celulas_unique)):
            
            wb = Workbook()
            wb = load_workbook('modelo_op_montagem.xlsx')
            ws = wb.active
            
            filtro_excel = (tab_completa['Célula'] == celulas_unique[0][i])
            filtrar = tab_completa.loc[filtro_excel]
            filtrar.reset_index(inplace=True)
            filtro_excel = (tab_completa['Célula'] == celulas_unique[0][i])
            
            if len(filtrar) > 21:
            
                for j in range(0,21):
                 
                    ws['G5'] = celulas_unique[0][i] # nome da coluna é '0'
                    ws['AD5'] = hoje #  data de hoje
                    ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico + celulas_unique[0][i][:4] #código único para cada sequenciamento
                    
                    if celulas_unique[0][i] == "EIXO COMPLETO":
                        ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico + "C" 
                            
                    if celulas_unique[0][i] == "EIXO SIMPLES":
                        ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico + "S"
                    
                    else:
                        ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico #código único para cada sequenciamento
                    
                    filtrar = tab_completa.loc[filtro_excel]
                    filtrar.reset_index(inplace=True)
                    
                    ws['M4']  = tipo_filtro  # data da carga
                    ws['B' + str(k)] = filtrar['Código'][j]
                    ws['G' + str(k)] = filtrar['Peca'][j]
                    ws['AD' + str(k)] = filtrar['Qtde_total'][j]
                    k = k + 1
                    
                    wb.template = False
                    wb.save('Montagem ' + celulas_unique[0][i] + '1.xlsx')
           
                my_file = "Montagem " + celulas_unique[0][i] +'1.xlsx'
                filenames.append(my_file)                
                
                k = 9
                
                wb = Workbook()
                wb = load_workbook('modelo_op_montagem.xlsx')
                ws = wb.active
        
                filtro_excel = (tab_completa['Célula'] == celulas_unique[0][i])
                filtrar = tab_completa.loc[filtro_excel]
                filtrar.reset_index(inplace=True)
                filtro_excel = (tab_completa['Célula'] == celulas_unique[0][i])
        
                if len(filtrar) > 21:

                    j = 21

                    for j in range(21,len(filtrar)):
                     
                        ws['G5'] = celulas_unique[0][i] # nome da coluna é '0'
                        ws['AD5'] = hoje #  data de hoje
                        
                        if celulas_unique[0][i] == "EIXO COMPLETO":
                            ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico + "C" 
                            
                        if celulas_unique[0][i] == "EIXO SIMPLES":
                            ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico + "S"
                        
                        else:
                            ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico #código único para cada sequenciamento
                                             
                        filtrar = tab_completa.loc[filtro_excel]
                        filtrar.reset_index(inplace=True)
                        
                        ws['M4']  = tipo_filtro  # data da carga
                        ws['B' + str(k)] = filtrar['Código'][j]
                        ws['G' + str(k)] = filtrar['Peca'][j]
                        ws['AD' + str(k)] = filtrar['Qtde_total'][j]
                        k = k + 1
                        
                        wb.template = False
                        wb.save('Montagem ' + celulas_unique[0][i] + '.xlsx')

                    my_file = "Montagem " + celulas_unique[0][i] +'.xlsx'
                    filenames.append(my_file)              

            else:
                
                j = 0
                k = 9
                    
                for j in range(0,21-(21-len(filtrar))):
                 
                    ws['G5'] = celulas_unique[0][i] # nome da coluna é '0'
                    ws['AD5'] = hoje #  data de hoje
                    
                    if celulas_unique[0][i] == "EIXO COMPLETO":
                        ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico + "C" 
                            
                    if celulas_unique[0][i] == "EIXO SIMPLES":
                        ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico + "S"            
                    
                    else:
                        
                        ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico #código único para cada sequenciamento
                    
                    filtrar = tab_completa.loc[filtro_excel]
                    filtrar.reset_index(inplace=True)
                    
                    ws['M4']  = tipo_filtro  # data da carga
                    ws['B' + str(k)] = filtrar['Código'][j]
                    ws['G' + str(k)] = filtrar['Peca'][j]
                    ws['AD' + str(k)] = filtrar['Qtde_total'][j]
                    k = k + 1
                    
                    wb.template = False
                    wb.save('Montagem ' + celulas_unique[0][i] + '.xlsx')
                
                k = 9
                    
                my_file = "Montagem " + celulas_unique[0][i] +'.xlsx'
                filenames.append(my_file)
               
    if setor == 'Solda':   
    
        #####colunas de códigos#####
        
        base_carretas['Código'] = base_carretas['Código'].astype(str) 
        base_carretas['Recurso'] = base_carretas['Recurso'].astype(str)
        
        #######retirando cores dos códigos######
        
        base_carga['Recurso']=base_carga['Recurso'].str.replace('AM','')
        base_carga['Recurso']=base_carga['Recurso'].str.replace('AN','')
        base_carga['Recurso']=base_carga['Recurso'].str.replace('VJ','')
        base_carga['Recurso']=base_carga['Recurso'].str.replace('LC','')
        base_carga['Recurso']=base_carga['Recurso'].str.replace('VM','')
        base_carga['Recurso']=base_carga['Recurso'].str.replace('AV','')
        
        ######retirando espaco em branco####
        
        base_carga['Recurso'] = base_carga['Recurso'].str.strip()
        
        #####excluindo colunas e linhas#####
        
        base_carretas.drop(['Etapa','Etapa2'], axis=1, inplace=True)
        
        base_carretas.drop(base_carretas[(base_carretas['Etapa3']  == '')].index, inplace=True) # & (base_carretas['Unit_Price'] < 600)].index, inplace=True)
    
        ####criando código único#####
        
        codigo_unico = tipo_filtro[:2] + tipo_filtro[3:5] + tipo_filtro[6:10]
        
        ####filtrando data da carga#####
        
        datas_unique = pd.DataFrame(base_carga['Datas'].unique())
        
        escolha_data = (base_carga['Datas'] == tipo_filtro)
        filtro_data = base_carga.loc[escolha_data]
        filtro_data['Datas'] = pd.to_datetime(filtro_data.Datas)
        
        #####juntando planilhas de acordo com o recurso#######
        
        tab_completa = pd.merge(filtro_data, base_carretas[['Recurso','Código','Peca','Qtde','Célula']], on=['Recurso'], how='left')
        tab_completa = tab_completa.dropna(axis=0)
        
        tab_completa['Código'] = tab_completa['Código'].astype(str)
        
        tab_completa.reset_index(inplace=True, drop=True)
        
        celulas_unique = pd.DataFrame(tab_completa['Célula'].unique())
        celulas_unique = celulas_unique.dropna(axis=0)
        celulas_unique.reset_index(inplace=True)
        
        recurso_unique =  pd.DataFrame(tab_completa['Recurso'].unique())
        recurso_unique = recurso_unique.dropna(axis=0)
                
        #criando coluna de quantidade total de itens
        
        try:
            tab_completa['Qtde_x'] = tab_completa['Qtde_x'].str.replace(',','.')
        except:
            pass
        
        tab_completa['Qtde_x'] = tab_completa['Qtde_x'].astype(float)
        tab_completa['Qtde_x'] = tab_completa['Qtde_x'].astype(int)
        
        tab_completa['Qtde_y'] = tab_completa['Qtde_y'].astype(float)
        tab_completa['Qtde_y'] = tab_completa['Qtde_y'].astype(int)
        
        tab_completa['Qtde_total'] = tab_completa['Qtde_x'] * tab_completa['Qtde_y']
        
        tab_completa = tab_completa.drop(columns=['Carga','Recurso','Qtde_x','Qtde_y'])
        
        tab_completa = tab_completa.groupby(['Código','Peca','Célula','Datas']).sum()
        
        #tab_completa = tab_completa.drop_duplicates()
        
        tab_completa.reset_index(inplace=True)
        
        ######tratando coluna de código e recurso
        
        for d in range(0,tab_completa.shape[0]):
        
            if len(tab_completa['Código'][d]) == 5:
                tab_completa['Código'][d] = '0' + tab_completa['Código'][d]
                
        #criando coluna de código para arquivar
        
        hoje = datetime.now()  
        
        ts = pd.Timestamp(hoje)
        
        hoje1 = hoje.strftime('%d%m%Y') #/
        
        controle_seq = tab_completa
        controle_seq["codigo"] = hoje1 + tipo_filtro
        
        st.write("Arquivos para download")    
        
        k = 9
        
        for i in range(0,len(celulas_unique)):
         
            wb = Workbook()
            wb = load_workbook('modelo_op_solda.xlsx')
            ws = wb.active
            
            filtro_excel = (tab_completa['Célula'] == celulas_unique[0][i])
            filtrar = tab_completa.loc[filtro_excel]
            filtrar.reset_index(inplace=True)
            filtro_excel = (tab_completa['Célula'] == celulas_unique[0][i])
            
            if len(filtrar) > 21:
            
                for j in range(0,21):
                 
                    ws['G5'] = celulas_unique[0][i] # nome da coluna é '0'
                    ws['AD5'] = hoje #  data de hoje
                    ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico + celulas_unique[0][i][:4] #código único para cada sequenciamento
                    
                    if celulas_unique[0][i] == "EIXO COMPLETO":
                        ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico + "C" 
                            
                    if celulas_unique[0][i] == "EIXO SIMPLES":
                        ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico + "S"
                    
                    else:
                        ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico #código único para cada sequenciamento
                    
                    filtrar = tab_completa.loc[filtro_excel]
                    filtrar.reset_index(inplace=True)
                    
                    ws['M4']  = tipo_filtro  # data da carga
                    ws['B' + str(k)] = filtrar['Código'][j]
                    ws['G' + str(k)] = filtrar['Peca'][j]
                    ws['AD' + str(k)] = filtrar['Qtde_total'][j]
                    k = k + 1
                    
                    wb.template = False
                    wb.save('Solda ' + celulas_unique[0][i] + '1.xlsx')
                
                my_file = "Solda " + celulas_unique[0][i] +'1.xlsx'
                filenames.append(my_file) 
                    
                k = 9
                
                wb = Workbook()
                wb = load_workbook('modelo_op_solda.xlsx')
                ws = wb.active
        
                filtro_excel = (tab_completa['Célula'] == celulas_unique[0][i])
                filtrar = tab_completa.loc[filtro_excel]
                filtrar.reset_index(inplace=True)
                filtro_excel = (tab_completa['Célula'] == celulas_unique[0][i])
        
                if len(filtrar) > 21:
                    
                    j = 21
                    
                    for j in range(21,len(filtrar)):
                     
                        ws['G5'] = celulas_unique[0][i] # nome da coluna é '0'
                        ws['AD5'] = hoje #  data de hoje
                        
                        if celulas_unique[0][i] == "EIXO COMPLETO":
                            ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico + "C" 
                            
                        if celulas_unique[0][i] == "EIXO SIMPLES":
                            ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico + "S"
                        
                        else:
                            ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico #código único para cada sequenciamento
                                             
                        filtrar = tab_completa.loc[filtro_excel]
                        filtrar.reset_index(inplace=True)
                        
                        ws['M4']  = tipo_filtro  # data da carga
                        ws['B' + str(k)] = filtrar['Código'][j]
                        ws['G' + str(k)] = filtrar['Peca'][j]
                        ws['AD' + str(k)] = filtrar['Qtde_total'][j]
                        k = k + 1
                        
                        wb.save('Solda ' + celulas_unique[0][i] + '.xlsx')
                
            else:
                
                j = 0
                k = 9
                for j in range(0,21-(21-len(filtrar))):
                 
                    ws['G5'] = celulas_unique[0][i] # nome da coluna é '0'
                    ws['AD5'] = hoje #  data de hoje
                    
                    if celulas_unique[0][i] == "EIXO COMPLETO":
                        ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico + "C" 
                            
                    if celulas_unique[0][i] == "EIXO SIMPLES":
                        ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico + "S"            
                    
                    else:
                        
                        ws['AK4'] = celulas_unique[0][i][0:3] + codigo_unico #código único para cada sequenciamento
                    
                    filtrar = tab_completa.loc[filtro_excel]
                    filtrar.reset_index(inplace=True)
                    
                    ws['M4']  = tipo_filtro  # data da carga
                    ws['B' + str(k)] = filtrar['Código'][j]
                    ws['G' + str(k)] = filtrar['Peca'][j]
                    ws['AD' + str(k)] = filtrar['Qtde_total'][j]
                    k = k + 1
                    
                    wb.template = False
                    wb.save('Solda ' + celulas_unique[0][i] + '.xlsx')
                    
                k = 9
                
                my_file = "Solda " + celulas_unique[0][i] +'.xlsx'

                filenames.append(my_file)
        
    filenames_unique = list(set(filenames))
    
    with zipfile.ZipFile("Arquivos.zip", mode="w") as archive:
        for filename in filenames_unique:
            archive.write(filename)
    
    with open("Arquivos.zip", "rb") as fp:
        btn = st.download_button(
            label="Download arquivos",
            data=fp,
            file_name="Arquivos.zip",
            mime="application/zip"
        )
            
    with st.sidebar:
    
        if att_controle == 'Atualizar':
            
            if setor != 'Selecione':
                    
                name_sheet2 = 'Controle de sequenciamento'
                worksheet4 = 'Controle'
                
                sh = sa.open(name_sheet2)
                
                lista_controle = {'Data': [today], 'Setor': [setor], 'Data_carga': [tipo_filtro]}
                
                controle_seq = pd.DataFrame(lista_controle)
                
                controle_seq = controle_seq.values.tolist()
                sh.values_append('Controle', {'valueInputOption': 'RAW'}, {'values': controle_seq})   

    st.write("Resumo:")
    base_carga_filtro = base_carga.query("Datas == @tipo_filtro")
    base_carga_filtro.dropna(inplace=True)
    base_carga_filtro = base_carga_filtro[['Recurso','Qtde']]
    base_carga_filtro['Qtde'] = base_carga_filtro['Qtde'].astype(int)
    base_carga_filtro = base_carga_filtro.groupby('Recurso').sum()
    
    tab_completa[['Célula','Código','Peca','Qtde_total']]      
